import streamlit as st
import sys
import os
import re
import json
import time
import tempfile
from io import BytesIO

# --- CONFIGURAÇÃO INICIAL (OBRIGATORIAMENTE O PRIMEIRO COMANDO ST) ---
st.set_page_config(page_title="Automação RAE CAIXA", page_icon="🏛️", layout="centered")

# --- BANCO DE DADOS DE PROFISSIONAIS ---
PROFISSIONAIS = {
    "FRANCISCO DAVID MENESES DOS SANTOS": {
        "empresa": "FRANCISCO DAVID MENESES DOS SANTOS - F. D. MENESES DOS SANTOS",
        "cnpj": "54.801.096/0001-16",
        "cpf_emp": "058.756.003-73",
        "nome_resp": "FRANCISCO DAVID MENESES DOS SANTOS",
        "cpf_resp": "058.756.003-73",
        "registro": "336241CE"
    },
    "PALLOMA TEIXEIRA DA SILVA": {
        "empresa": "PALLOMA TEIXEIRA DA SILVA - PALLOMA TEIXEIRA ARQUITETURA LTDA",
        "cnpj": "54.862.474/0001-71",
        "cpf_emp": "064.943.593-10",
        "nome_resp": "PALLOMA TEIXEIRA DA SILVA",
        "cpf_resp": "064.943.593-10",
        "registro": "A184355-9"
    },
    "SANDY PEREIRA CORDEIRO": {
        "empresa": "SANDY PEREIRA CORDEIRO - CS ENGENHARIA",
        "cnpj": "54.794.898/0001-46",
        "cpf_emp": "071.222.553-60",
        "nome_resp": "SANDY PEREIRA CORDEIRO",
        "cpf_resp": "071.222.553-60",
        "registro": "356882CE"
    },
    "TIAGO VICTOR DE SOUSA": {
        "empresa": "TIAGO VICTOR DE SOUSA - T V S ENGENHARIA E ASSESSORIA",
        "cnpj": "54.806.521/0001-60",
        "cpf_emp": "068.594.803-00",
        "nome_resp": "TIAGO VICTOR DE SOUSA",
        "cpf_resp": "068.594.803-00",
        "registro": "346856CE"
    }
}

# --- PATCH DE METADADOS ULTRA-ROBUSTO ---
try:
    import importlib.metadata as metadata
except ImportError:
    import importlib_metadata as metadata

_original_version = metadata.version
def patched_version(package_name):
    try:
        return _original_version(package_name)
    except Exception:
        versions = {
            'docling': '2.15.0',
            'docling-core': '2.9.0',
            'docling-parse': '2.4.0',
            'docling-ibm-models': '1.1.0',
            'pypdfium2': '4.30.0',
            'openpyxl': '3.1.5',
            'transformers': '4.40.0',
            'torch': '2.2.0',
            'torchvision': '0.17.0',
            'timm': '0.9.16',
            'optree': '0.11.0'
        }
        return versions.get(package_name, "1.0.0")
metadata.version = patched_version

# --- IMPORTAÇÃO DAS DEPENDÊNCIAS ---
try:
    import pandas as pd
    from openpyxl import load_workbook
    from docling.document_converter import DocumentConverter, PdfFormatOption
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.datamodel.base_models import InputFormat
    import google.generativeai as genai
    import onnxruntime
    import transformers
    import timm
    import optree 
    DEPENDENCIAS_OK = True
    ERRO_IMPORT = ""
except ImportError as e:
    DEPENDENCIAS_OK = False
    ERRO_IMPORT = str(e)

# --- ESTILIZAÇÃO ---
st.markdown("""
    <style>
    .main { background-color: #ffffff; }
    .stButton>button {
        width: 100%; border-radius: 8px; height: 3.5em;
        background-color: #4f46e5; color: white; font-weight: bold; border: none;
    }
    .stDownloadButton>button {
        width: 100%; border-radius: 8px;
        background-color: #059669; color: white; border: none;
    }
    </style>
    """, unsafe_allow_html=True)

# Cache para o conversor
@st.cache_resource
def get_converter():
    pipeline_options = PdfPipelineOptions()
    pipeline_options.do_table_structure = True 
    pipeline_options.table_structure_options.do_cell_matching = True
    return DocumentConverter(
        allowed_formats=[InputFormat.PDF],
        format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options)
        }
    )

def call_gemini(api_key, prompt):
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel('gemini-2.5-flash')
    for attempt in range(3):
        try:
            response = model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    response_mime_type="application/json",
                    temperature=0.1
                )
            )
            return json.loads(response.text)
        except Exception as e:
            if attempt == 2: raise e
            time.sleep(2)

def main():
    st.title("🏛️ Automação RAE CAIXA")
    st.markdown("##### Inteligência Artificial para Engenharia")

    if not DEPENDENCIAS_OK:
        st.error(f"❌ Erro de Dependências: {ERRO_IMPORT}")
        return

    with st.sidebar:
        st.header("⚙️ Configurações")
        api_key = st.text_input("Gemini API Key:", type="password")
        
        st.divider()
        st.subheader("👤 Responsável Técnico")
        resp_selecionado = st.selectbox(
            "Selecione o Profissional:",
            options=list(PROFISSIONAIS.keys())
        )
        st.divider()
        st.caption("v3.7 - Seleção de Profissionais")

    col1, col2 = st.columns(2)
    with col1:
        pdf_file = st.file_uploader("1. Enviar Laudo (PDF)", type=["pdf"])
    with col2:
        excel_file = st.file_uploader("2. Enviar Modelo (.xlsm)", type=["xlsm"])

    if st.button("🚀 INICIAR PROCESSAMENTO"):
        if not api_key or not pdf_file or not excel_file:
            st.warning("Preencha a chave e carregue os ficheiros.")
            return

        try:
            with st.status("A processar laudo técnico...", expanded=True) as status:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                    tmp.write(pdf_file.getbuffer())
                    tmp_path = tmp.name

                try:
                    st.write("📖 Lendo estrutura do PDF com Docling...")
                    converter = get_converter()
                    res = converter.convert(tmp_path)
                    md_content = re.sub(r'\n\s*\n', '\n', res.document.export_to_markdown())
                finally:
                    if os.path.exists(tmp_path): os.remove(tmp_path)

                st.write("🧠 IA: Extraindo dados técnicos...")
                prompt = f"""
                Atue como engenheiro revisor da CAIXA. Extraia os dados para JSON:
                - CAMPOS: proponente, cpf_cnpj, ddd, telefone, endereco, bairro, cep, municipio, uf_vistoria, uf_registro, complemento, matricula, comarca, valor_terreno, valor_imovel, lat_s, long_w, etapas_original
                - OFICIO: Número após a matrícula em DOCUMENTOS (ex: 12345 / 3 / CE, ofício é 3).
                - COORDENADAS (GMS puro): 
                    - lat_s: Latitude (ex: 06°24'08.8"). NÃO inclua letras (S/N).
                    - long_w: Longitude (ex: 39°18'21.5"). NÃO inclua letras (W/E).
                - CRONOGRAMA: etapas_original (Número total de etapas/meses).
                - TABELAS: 'incidencias' (20 números PESO %), 'acumulado' (percentuais % ACUMULADO).
                DOCUMENTO: {md_content}
                """
                dados = call_gemini(api_key, prompt)

                st.write("📊 Gravando na planilha Excel...")
                wb = load_workbook(BytesIO(excel_file.read()), keep_vba=True)
                wb.calculation.fullCalcOnLoad = True

                def to_f(v):
                    if isinstance(v, (int, float)): return v
                    try: return float(str(v).replace(',', '.').replace('%', '').strip())
                    except: return 0

                # Aba Início Vistoria
                if "Início Vistoria" in wb.sheetnames:
                    ws = wb["Início Vistoria"]
                    mapping = {
                        "G43": "proponente", "AJ43": "cpf_cnpj", "AP43": "ddd", "AR43": "telefone",
                        "G49": "endereco", "AD49": "lat_s", "AH49": "long_w", "AL49": "complemento",
                        "G51": "bairro", "V51": "cep", "AA51": "municipio", "AS51": "uf_vistoria",
                        "AS53": "uf_registro", "G53": "valor_terreno", "Q53": "matricula",
                        "AA53": "oficio", "AJ53": "comarca"
                    }
                    for cell, key in mapping.items():
                        val = dados.get(key, "")
                        ws[cell] = to_f(val) if key == "valor_terreno" else str(val).upper()
                    ws["Q54"], ws["Q55"], ws["Q56"] = "Casa", "Residencial", "Vistoria para aferição de obra"

                # Aba RAE
                if "RAE" in wb.sheetnames:
                    ws_rae = wb["RAE"]
                    ws_rae.sheet_state = 'visible'
                    ws_rae["AH66"] = to_f(dados.get("valor_imovel", 0))
                    ws_rae["AS66"] = to_f(dados.get("etapas_original", 0))
                    
                    # Preenchimento do Profissional Selecionado
                    prof = PROFISSIONAIS[resp_selecionado]
                    ws_rae["I315"] = prof["empresa"].upper()
                    ws_rae["I316"] = prof["cnpj"]
                    ws_rae["U316"] = prof["cpf_emp"]
                    ws_rae["AE315"] = prof["nome_resp"].upper()
                    ws_rae["AE316"] = prof["cpf_resp"]
                    ws_rae["AO316"] = prof["registro"].upper()
                    
                    incs, acus = dados.get("incidencias", []), dados.get("acumulado", [])
                    for i in range(20):
                        ws_rae[f"S{69+i}"] = to_f(incs[i]) if i < len(incs) else 0
                    for i in range(len(acus)):
                        if i < 37: ws_rae[f"AE{72+i}"] = to_f(acus[i])

                output = BytesIO()
                wb.save(output)
                processed_data = output.getvalue()
                
                proponente = dados.get("proponente", "").strip()
                primeiro_nome = proponente.split(' ')[0].upper() if proponente else "FINAL"
                nome_arq = f"RAE_{primeiro_nome}.xlsm"

                status.update(label="✅ Mapeamento concluído!", state="complete", expanded=False)

            st.balloons()
            st.download_button(
                label=f"📥 BAIXAR RAE - {primeiro_nome}",
                data=processed_data,
                file_name=nome_arq,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12"
            )

        except Exception as e:
            st.error(f"Erro no processamento: {e}")

if __name__ == "__main__":
    main()
